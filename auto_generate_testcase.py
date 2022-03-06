#!/usr/bin/env python
# -*- encoding: utf-8 -*-
"""
@author: Stone
@license: (C) Copyright 2017-2022, Node Supply Chain Manager Corporation Limited.
@contact: leicheng2013@outlook.com
@software: ChinaSoft
@file: auto_generate_testcase.py
@time: 2022/2/18 20:24
@desc:
"""

from xmindparser import xmind_to_dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
from openpyxl.utils import get_column_letter
import os


class AutoGenerateTestcase:

    def __init__(self, data):
        self.sub_system = data['sub_system']
        self.um_code = data['um_code']
        self.is_trans = data['is_trans']
        self.significance = data['significance']
        self.testcases = []

    def __parse_xmind(self, filepath):
        print('开始解析xmind文件')
        return xmind_to_dict(file_path=filepath)[0]['topic']

    def __parse_origin_data(self, datadict, cur='', excel=None):
        '''
        解析xmind转换的原始数据，生成所有测试用需要的关键信息
        :param datadict:
        :param cur:
        :param excel:
        :return:
        '''
        for k, v in datadict.items():
            if type(v) is dict:
                self.__parse_origin_data(v, cur, excel)
            elif isinstance(v, list):
                for i in v:
                    if isinstance(i, dict):
                        self.__parse_origin_data(i, cur, excel)
            else:
                if 'topics' not in datadict and k == 'title':
                    if '改造点' not in cur:
                        case_info = (cur + '-' + v).split('-')[1:]
                        case = self.__package_testcase(case_info)
                        return self.testcases.append(case)
                cur = cur + '-' + v

    def __package_testcase(self, data: list):
        '''
        将测试用例的信息排列组合成测试用例，并输出为列表
        :param data:
        :return:
        '''
        case_import_path = '-'.join(data[1:3])  # 导入路径随便填，需要手动修改
        testcase_name = '-'.join(data[2:-2]).replace('-测试点', '').replace('-正向', '').replace('-反向', '')  # 去除用例名称中的测试点等字样
        expect_result = data[-2]
        if '正向' in data[-4]:
            direction = '正例'
        else:
            direction = '反例'
        level = data[-1]
        if '接口' not in data[2]:  # 功能模块中含有接口字样，这判断为接口的用例，用例步骤描述稍有区别
            step1 = '1、用户进入"' + '-->'.join(data[2:4]) + '"界面\n'
            step2 = '2、' + data[-3]
            step_desc = step1 + step2
        else:
            step1 = '1、使用postman调用接口'
            step_desc = step1
        day = datetime.now().strftime('%Y-%m-%d')
        case_li = [case_import_path, self.sub_system, testcase_name, '', step_desc, expect_result, direction, level,
                   self.is_trans, self.um_code, day, '', '', '', self.significance]
        return case_li

    def save_to_excel(self, testcases: list):
        '''
        将测试用例写入excel文件
        :param testcases:
        :return:
        '''
        print('开始写入excel文件。。。')
        if not os.path.exists(file_of_testcase):  # 判断用例文件是否存在，不存在就新建，存在就加载并修改
            wb = Workbook()
            ws = wb.active
            # 写入第一行表头
            ws.append(['案例导入路径', '子系统', '测试案例名称', '案例描述', '步骤描述', ''])

            for testcase in testcases:
                ws.append(testcase)
                wb.save(filename=file_of_testcase)
                wb.close()
        else:
            wb = load_workbook(filename=file_of_testcase)
            ws = wb.active
            row = 2  # 文件已存在，只需要从第二行开始覆盖数据即可
            for testcase in testcases:
                for i in range(1, len(testcase)):
                    ws.cell(row=row, column=i, value=testcase[i - 1])
                row += 1
            wb.save(filename=file_of_testcase)
            wb.close()
        print('测试用例文件写入完成，开始美化。。。')
        self.__beauty_excel()

    def __beauty_excel(self):
        '''
        美化生成的测试用例，设置行高，列宽，背景色以及边框
        :return:
        '''
        wb = load_workbook(filename=file_of_testcase)
        ws = wb.active
        # 定义颜色
        orange = 'FF9933'
        green = 'CCFFCC'
        # 边框样式
        boder = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                       top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
        # 调整第一行行高
        ws.row_dimensions[1].height = 40
        # 分别设置列宽
        for col in ws.columns:
            index = list(ws.columns).index(col)  # 获取序列号
            letter = get_column_letter(index + 1)  # 获取序列号对应的字母
            if letter in ['A', 'C', 'E', 'F']:
                ws.column_dimensions[letter].width = 36
            elif letter == 'B':
                ws.column_dimensions[letter].width = 20
            else:
                ws.column_dimensions[letter].width = 12
        # 设置表格垂直居中，除第一行外奇数行填充浅绿色背景色以及设置边框
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if row == 1:  # 设置第一行的字体，背景色以及对齐方式
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center',
                                                                       wrap_text=True)
                    ws.cell(row=row, column=col).fill = PatternFill(fgColor=orange, fill_type='solid')
                    ws.cell(row=row, column=col).font = Font(size=12, b=True)
                    # b为True表示加粗、大小为12
                elif row >= 2 and col >= 7:  # 设置后面几列的对齐方式
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center',
                                                                       wrap_text=True)
                else:
                    ws.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
                ws.cell(row=row, column=col).border = boder  # 统一设置全边框

                if row >= 2 and row % 2:
                    ws.cell(row=row, column=col).fill = PatternFill(fgColor=green, fill_type='solid')
        wb.save(filename=file_of_testcase)
        wb.close()
        print('测试用例美化完成！')

    def generate_testcase(self, filepath):
        start_time = datetime.now().timestamp()
        origin_data = self.__parse_xmind(filepath)
        print('解析原始数据，获取测试用例关键信息，请稍后。。。')
        self.__parse_origin_data(origin_data)
        self.save_to_excel(self.testcases)
        end_time = datetime.now().timestamp()
        print('此次自动生成测试用例任务已完成，共耗时{time:.3f}秒。'.format(time=end_time - start_time))


if __name__ == '__main__':
    xmind_filepath = ''
    info = {
        "author": "author",
        "um_code": "umCode",
        "sub_system": "BOBS-OBP-开放银行",
        "is_trans": "否",
        "significance": "1"
    }
    # file_of_testcase = r''
    file_of_testcase = xmind_filepath.split('-')[0] + '测试用例-{}.xlsx'.format(info['author'])
    auto = AutoGenerateTestcase(info)
    auto.generate_testcase(xmind_filepath)
